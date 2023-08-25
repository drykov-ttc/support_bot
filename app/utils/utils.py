from datetime import datetime
import logging
import os
from openpyxl import load_workbook
import requests
from tabulate import tabulate
from prettytable import PrettyTable
import config
from config import ROOT_DIR
import calendar

logger = logging.getLogger(__name__)
raspDir = os.path.join(ROOT_DIR, "Rasp/")
curr_month = datetime.now().strftime("%m.%Y")

fileNAme = os.path.join(raspDir, curr_month + ".xlsx")


def get_month():
    months_russian = [
        "январь",
        "февраль",
        "март",
        "апрель",
        "май",
        "июнь",
        "июль",
        "август",
        "сентябрь",
        "октябрь",
        "ноябрь",
        "декабрь",
    ]
    current_month = datetime.now().month
    month = months_russian[current_month - 1]
    return month


def get_people():
    now = datetime.today().day
    last = 0
    count_people = 0
    people = []
    nicknames = []
    wb = load_workbook(fileNAme)
    print(wb)
    sheet = wb["Лист1"]
    for i in range(3, 10):
        if (
            sheet.cell(row=i, column=(now + 5)).value == "Х"
            or sheet.cell(row=i, column=(now + 5)).value == 1
        ):
            today_id = i
    while last != 1:
        if sheet.cell(row=3 + count_people, column=(1)).value != None:
            people.append(sheet.cell(row=3 + count_people, column=(1)).value)
            nicknames.append(sheet.cell(row=3 + count_people, column=(2)).value)
            count_people += 1
        else:
            last = 1
    print(f"{people, count_people, nicknames, today_id}")
    return (people, count_people, nicknames, today_id)
    wb.save(fileNAme)


def main_duty_new(now, q):
    try:
        data = []
        wb = load_workbook(fileNAme)
        sheet = wb["Лист1"]
        for i in range(3, 10):
            if (
                sheet.cell(row=i, column=(now + 5)).value == "Х"
                or sheet.cell(row=i, column=(now + 5)).value == 1
            ):
                data.append(["", sheet.cell(row=i, column=3).value])
                data.append(
                    [
                        sheet.cell(row=i, column=1).value,
                        sheet.cell(row=i, column=5).value,
                    ]
                )
        table_month = sheet.cell(row=1, column=2).value
        table = tabulate(
            data, tablefmt="grid", stralign="center"
        )  # You can use different table formats
        return (table, table_month)
    except Exception as e:
        logger.error(f"{e}")
        print(f"Error: {e}")
