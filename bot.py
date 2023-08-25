# -*- coding: utf-8 -*-
import logging
import config
import telebot
import time
import requests
from telebot import types, apihelper

from datetime import datetime
import threading
import os
from openpyxl import load_workbook
import getpass


logFile = os.path.join(config.WORK_DIR, config.LOG_DIR, config.LOG_STABLE)
picDir = os.path.join(config.WORK_DIR, "pic/")
raspDir = os.path.join(config.WORK_DIR, "Rasp/")
current_data = datetime.now()
curr_month = current_data.strftime("%m.%Y")
wb = load_workbook(os.path.join(raspDir, curr_month + ".xlsx"))
fileNAme = os.path.join(raspDir, curr_month + ".xlsx")
bot = telebot.TeleBot(config.token)

USER_NAME = getpass.getuser()


logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
    filename=logFile,
)

logger = logging.getLogger(__name__)


def add_to_startup(file_path=""):
    if file_path == "":
        file_path = os.path.dirname(os.path.realpath(__file__))
    bat_path = (
        r"C:\Users\%s\AppData\Roaming\Microsoft\Windows\Start Menu\Programs\Startup"
        % USER_NAME
    )
    with open(bat_path + "\\" + "open.bat", "w+") as bat_file:
        bat_file.write(r'start "" %s' % file_path)


@bot.message_handler(commands=["help"])  # Узнать список команд
def help(m):
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add(
        *[
            types.KeyboardButton(name)
            for name in ["Дежурный сегодня?", "Дежурный завтра?"]
        ]
    )
    bot.send_message(
        m.chat.id,
        "Список команд: \n/today - Дежурный сегодня \n/tomorrow - Дежурный завтра "
        "\n/duty - График дежурств \n/change_duty - Смена дежурного на сегодня \n/all - "
        "Сделать оповещение",
        reply_markup=keyboard,
    )


@bot.message_handler(commands=["duty"])
def find_file_ids(message):
    try:
        for file in os.listdir(picDir):
            if file.split(".")[-1] == "png":
                picFile = os.path.join(picDir, file)
                f = open(picFile, "rb")
                bot.send_photo(
                    message.chat.id, f, None, reply_markup=types.ReplyKeyboardRemove()
                )
            time.sleep(1)
    except Exception as e:
        logger.error(f"{e}")
        print(f"Error: {e}")
        bot.send_message(config.admin_id, e)


@bot.message_handler(commands=["start"])  # Приветствие + запуск клавиатуры
def start(m):
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add(
        *[
            types.KeyboardButton(name)
            for name in ["Дежурный сегодня?", "Дежурный завтра?"]
        ]
    )
    bot.send_message(m.chat.id, "Приветствую!", reply_markup=keyboard)


@bot.message_handler(commands=["Message"])
def marat():
    bot.send_message(
        config.chat_id,
        "Привет, коллеги!\n"
        "Небольшое обновление:\n\n"
        "✔ Переезд на новый сервер "
        "\n✔ Автозапуск бота и его прокси",
    )


@bot.message_handler(commands=["all"])
def msg_and(m):
    nicknames = [
        "@supervufel",
        "@white_yarkiy",
        "@Somtawl",
        "@Blackvvolf",
        "@mikhail_lavrenov",
        "@BorisBeloglazov",
        "@M4r4t1n4",
        "@Yok01337",
    ]
    message = "Важное оповещение!\n\n"
    formatted_nicknames = "\n".join(nicknames)

    bot.send_message(config.chat_id, message + formatted_nicknames)


def get_people():
    now = datetime.today().day
    last = 0
    count_people = 0
    people = []
    nicknames = []
    # wb = load_workbook("./Rasp/rasp.xlsx")
    sheet = wb["Лист1"]
    for i in range(3,10):
        if sheet.cell(row=i, column=(now + 5)).value == "Х":
            today_id = i
    while last != 1:
        if sheet.cell(row=3 + count_people, column=(1)).value != None:
            people.append(sheet.cell(row=3 + count_people, column=(1)).value)
            nicknames.append(sheet.cell(row=3 + count_people, column=(2)).value)
            count_people += 1
        else:
            last = 1
    print(f"{people, count_people, nicknames, today_id}")
    return (people, count_people, nicknames, today_id)
    wb.save(fileNAme)


@bot.message_handler(commands=["change_duty"])  # Смена дежурного
def change_duty(m):
    a = ["Oritorius"]
    ids = [632646448]
    try:
        if m.from_user.username in a and m.from_user.id in ids:
            a = get_people()
            people = a[0]
            count_people = a[1]
            nicknames = a[2]
            keyboard = types.InlineKeyboardMarkup()
            for i in range(0, count_people):
                nicknames[i] = types.InlineKeyboardButton(
                    text=people[i], callback_data=nicknames[i]
                )
                keyboard.add(nicknames[i])
            bot.send_message(
                m.chat.id, "Назначьте дежурного на сегодня:", reply_markup=keyboard
            )

    except Exception as e:
        logger.error(f"{e}")
        print(f"Error: {e}")
        bot.send_message(config.admin_id, e)


@bot.callback_query_handler(func=lambda call: True)
def callback_inline(call):
    try:
        # wb = load_workbook("./Rasp/rasp.xlsx")
        sheet = wb["Лист1"]
        now = datetime.today().day
        a = get_people()
        people = a[0]
        count_people = a[1]
        nicknames = a[2]
        today_id = a[3]
        for i in range(3,10):  # удаляем за сегодня
            sheet.cell(row=i, column=(now + 5)).value = None
        d = nicknames.index(call.data)  # узнаем кто будет сегодня дежурить
        next_column = 45
        for i in range(now + 3, 40):
            if sheet.cell(row=d + 3, column=(i)).value == "Х":
                next_column = i
                break
        for i in range(3,10):
            sheet.cell(row=i, column=next_column).value = None
        sheet.cell(row=today_id, column=next_column).value ="Х"
        sheet.cell(row=d + 3, column=(now + 5)).value = "Х"
        bot.send_message(
            call.message.chat.id,
            people[d] + " " + nicknames[d] + " назначен дежурным на сегодня",
            reply_markup=types.ReplyKeyboardRemove(),
        )
        bot.answer_callback_query(
            callback_query_id=call.id,
            show_alert=True,
            text="Вы изменили дежурного на сегодня!",
        )
        bot.delete_message(
            call.message.chat.id,
            call.message.message_id,
        )
        wb.save(fileNAme)
    except Exception as e:
        logger.error(f"{e}")
        print(f"Error: {e}")
        bot.send_message(config.admin_id, e)


@bot.message_handler(commands=["today"])  # Узнать кто дежурит сегодня
def today(m):
    try:
        q = 0
        now = datetime.today().day
        q = main_duty_new(now, q)
        q = q[0]
        if q != 0:
            bot.send_message(
                m.chat.id,
                "Сегодня дежурит\n" + q,
                reply_markup=types.ReplyKeyboardRemove(),
            )
        else:  # исключение для того, чтобы был фидбэк пользователю
            bot.send_message(
                m.chat.id,
                "Сегодня выходной! \nОтдохните от работы, погуляйте на свежем воздухе :)",
                reply_markup=types.ReplyKeyboardRemove(),
            )
    except Exception as e:
        logger.error(f"{e}")
        print(f"Error: {e}")
        bot.reply_to(m, "Ошибочка!", reply_markup=types.ReplyKeyboardRemove())
        bot.send_message(config.admin_id, e)


@bot.message_handler(commands=["tomorrow"])  # Узнать кто дежурит завтра
def tomorrow(m):
    try:
        q = 0
        now = datetime.today().day + 1
        q = main_duty_new(now, q)
        q = q[0]
        if q != 0:
            bot.send_message(
                m.chat.id,
                "Завтра дежурит\n" + q,
                reply_markup=types.ReplyKeyboardRemove(),
            )
        else:
            bot.send_message(
                m.chat.id,
                "Завтра выходной! \nПроведите это время с пользой для себя :)",
                reply_markup=types.ReplyKeyboardRemove(),
            )
    except Exception as e:
        logger.error(f"{e}")
        print(f"Error: {e}")
        bot.reply_to(m, "Ошибочка!")
        bot.send_message(config.admin_id, e)


@bot.message_handler(content_types=["document"])
def handle_file(message):
    a = ["Oritorius"]
    ids = [632646448]
    try:
        if message.from_user.username in a and message.from_user.id in ids:
            file = requests.get(
                "https://api.telegram.org/bot{0}/getFile?file_id={1}".format(
                    config.token, message.document.file_id
                )
            )
            file_path = file.json()["result"]["file_path"]
            src = raspDir + message.document.file_name
            src2 = picDir + message.document.file_name
            url = "https://api.telegram.org/file/bot{0}/{1}".format(
                config.token, file_path
            )
            r = requests.get(url)
            if message.document.file_name == curr_month + ".xlsx":
                with open(src, "wb") as new_file:
                    new_file.write(r.content)
                bot.reply_to(message, "Новое расписание успешно загружено")
            elif message.document.file_name == curr_month + ".png":
                with open(src2, "wb") as new_file:
                    new_file.write(r.content)
                bot.reply_to(message, "Новая картинка расписания успешно загружено")
    except Exception as e:
        logger.error(f"{e}")
        print(f"Error: {e}")
        bot.reply_to(message, e)


def main_duty_new(now, q):
    try:
        duty_list = []
        # wb = load_workbook("./Rasp/rasp.xlsx")
        sheet = wb["Лист1"]
        for i in range(3, 10):
            if sheet.cell(row=i, column=(now + 5)).value == "Х":
                duty_list.append(
                    (sheet.cell(row=i, column=3).value)
                    + ": \n"
                    + (sheet.cell(row=i, column=1).value)
                    + "\n"
                    + (sheet.cell(row=i, column=5).value)+ "\n\n"
                ) 
                print(duty_list)
        table_month = sheet.cell(row=1, column=1).value
        print(duty_list)
        q = "\n".join(duty_list)
        print(f"{q}")
        return (q, table_month)
    except Exception as e:
        logger.error(f"{e}")
        print(f"Error: {e}")
        bot.send_message(config.admin_id, e)


@bot.message_handler(func=lambda message: True, content_types=["text"])
def duty_main(m):
    try:
        if m.text == "Дежурный сегодня?":
            today(m)
        elif m.text == "Дежурный завтра?":
            tomorrow(m)
        else:
            bot.send_message(
                m.chat.id,
                "К сожалению я не знаю данную команду. "
                "\nПопробуйте ввести /help, чтобы узнать доступные команды",
                reply_markup=types.ReplyKeyboardRemove(),
            )
    except Exception as e:
        logger.error(f"{e}")
        print(f"Error: {e}")
        bot.send_message(config.admin_id, e)


def clock(interval):  # Ежедневный постинг
    try:
        time_start1 = "08:50"
        while True:
            d = datetime.today()
            time_x = d.strftime("%H:%M")
            if time_x in time_start1:
                q = 0
                now = datetime.today().day
                month = datetime.today().month
                all = main_duty_new(now, q)
                q = all[0]
                table_month = all[1]
                print(table_month)
                print(month)
                if table_month == month:
                    if q != 0:
                        bot.send_message(
                            config.chat_id,
                            "Доброе утро, коллеги! \nСегодня дежурит " + q,
                        )
                        for file in os.listdir("pic/"):
                            if file.split(".")[-1] == "png":
                                f = open("pic/" + file, "rb")
                                bot.send_photo(
                                    config.chat_id,
                                    f,
                                    None,
                                    reply_markup=types.ReplyKeyboardRemove(),
                                )
                            time.sleep(1)
                else:
                    bot.send_message(
                        config.chat_id,
                        "Отсутсвует актуальное расписание на данный месяц!",
                    )

            time.sleep(interval)
    except Exception as e:
        logger.error(f"{e}")
        print(f"Error: {e}")
        bot.send_message(config.admin_id, e)


try:
    t = threading.Thread(target=clock, args=(60,))
    t.daemon = True
    t.start()
except Exception as e:
    logger.error(f"{e}")
    print(f"Error: {e}")
    bot.send_message(config.admin_id, e)

while True:
    try:
        if __name__ == "__main__":
            # bot.polling(none_stop=True)
            bot.polling(none_stop=True)

    except Exception as e:
        logger.error(f"{e}")
        print(f"Error: {e}")

        bot.send_message(config.admin_id, "Бот упал")
        time.sleep(15)
        bot.send_message(config.admin_id, "Alive!")
